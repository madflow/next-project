# pyright: reportMissingImports=false

from copy import copy
from dataclasses import dataclass, field
from importlib import import_module
from io import BytesIO
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText as ChartRichText
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.drawing.text import (
    Font as DrawingFont,
)
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles import Font as CellFont
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import cm_to_EMU
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DEFAULT_LIGHT_TEXT_HEX = "#FFFFFF"
DEFAULT_DARK_TEXT_HEX = "#000000"
DEFAULT_HEADER_FILL_HEX = "#3B82F6"
HEADER_ROW_INDEX = 4
DATA_ROW_INDEX = 5
TABLE_STYLE_NAME = "TableStyleMedium9"
EXPORT_FONT_NAME = "Arial"
SANS_SERIF_FONT_FAMILY = 2
INVALID_SHEET_TITLE_CHARS = "[]:*?/\\"
PERCENT_NUMBER_FORMAT = '0.00"%"'
STACKED_CHART_NUMBER_FORMAT = '0"%"'
PIE_PERCENT_NUMBER_FORMAT = "0%"
MEAN_NUMBER_FORMAT = "0.0"
CHARTSHEET_SIZE_MULTIPLIER = 2
CHART_BORDER_WIDTH = 9360
AXIS_BORDER_WIDTH = 3600
CHART_BACKGROUND_HEX = "#FFFFFF"
CHART_LINE_HEX = "#999999"


@dataclass
class WorksheetTableSpec:
    headers: list[str]
    rows: list[list[Any]]
    header_fill_colors: dict[int, str] = field(default_factory=dict)
    cell_fill_colors: dict[tuple[int, int], str] = field(default_factory=dict)
    number_formats: dict[int, str] = field(default_factory=dict)
    table_name: str = "ChartExportData"


def _normalize_hex_color(color: str) -> str:
    normalized = color.strip().removeprefix("#")
    if len(normalized) != 6:
        raise ValueError(f"Unsupported color value '{color}'")
    return normalized.upper()


def _argb_hex_color(color: str) -> str:
    return f"FF{_normalize_hex_color(color)}"


def _rgb_components(color: str) -> tuple[int, int, int]:
    normalized = _normalize_hex_color(color)
    return (
        int(normalized[0:2], 16),
        int(normalized[2:4], 16),
        int(normalized[4:6], 16),
    )


def _relative_luminance(color: str) -> float:
    def to_linear(channel: int) -> float:
        value = channel / 255
        if value <= 0.03928:
            return value / 12.92
        return ((value + 0.055) / 1.055) ** 2.4

    red, green, blue = _rgb_components(color)
    return (
        0.2126 * to_linear(red) + 0.7152 * to_linear(green) + 0.0722 * to_linear(blue)
    )


def _contrast_ratio(first_color: str, second_color: str) -> float:
    first_luminance = _relative_luminance(first_color)
    second_luminance = _relative_luminance(second_color)
    lighter = max(first_luminance, second_luminance)
    darker = min(first_luminance, second_luminance)
    return (lighter + 0.05) / (darker + 0.05)


def _best_text_color(background_color: str | None) -> str:
    if not background_color:
        return _argb_hex_color(DEFAULT_LIGHT_TEXT_HEX)

    try:
        dark_contrast = _contrast_ratio(background_color, DEFAULT_DARK_TEXT_HEX)
        light_contrast = _contrast_ratio(background_color, DEFAULT_LIGHT_TEXT_HEX)
    except ValueError:
        return _argb_hex_color(DEFAULT_LIGHT_TEXT_HEX)

    if dark_contrast >= light_contrast:
        return _argb_hex_color(DEFAULT_DARK_TEXT_HEX)

    return _argb_hex_color(DEFAULT_LIGHT_TEXT_HEX)


def _font_with_name(font: CellFont, *, font_name: str) -> CellFont:
    updated_font = copy(font)
    updated_font.name = font_name
    updated_font.family = SANS_SERIF_FONT_FAMILY
    updated_font.scheme = None
    return updated_font


def _chart_typeface(font_name: str) -> DrawingFont:
    return DrawingFont(typeface=font_name)


def _chart_character_properties(font_name: str) -> CharacterProperties:
    return CharacterProperties(
        latin=_chart_typeface(font_name),
        ea=_chart_typeface(font_name),
        cs=_chart_typeface(font_name),
        sym=_chart_typeface(font_name),
    )


def _chart_text_properties(font_name: str) -> ChartRichText:
    character_properties = _chart_character_properties(font_name)
    return ChartRichText(
        bodyPr=RichTextProperties(),
        p=[
            Paragraph(
                pPr=ParagraphProperties(defRPr=character_properties),
                endParaRPr=copy(character_properties),
            )
        ],
    )


def _set_workbook_default_font(workbook: Workbook, *, font_name: str) -> None:
    workbook_internal = cast(Any, workbook)
    fonts = cast(Any, workbook_internal._fonts)  # noqa: SLF001
    fonts[0] = _font_with_name(fonts[0], font_name=font_name)

    named_styles = cast(Any, workbook_internal._named_styles)  # noqa: SLF001
    named_styles["Normal"].font = copy(fonts[0])


def _apply_chart_title_font(chart: BarChart | PieChart, *, font_name: str) -> None:
    title = chart.title
    if title is None or title.tx is None or title.tx.rich is None:
        return

    title.txPr = _chart_text_properties(font_name)
    for paragraph in title.tx.rich.p:
        character_properties = _chart_character_properties(font_name)
        if paragraph.pPr is None:
            paragraph.pPr = ParagraphProperties()
        paragraph.pPr.defRPr = character_properties
        paragraph.endParaRPr = copy(character_properties)

        for run in paragraph.r or []:
            run.rPr = copy(character_properties)


def _apply_chart_font(chart: BarChart | PieChart, *, font_name: str) -> None:
    _apply_chart_title_font(chart, font_name=font_name)

    legend = chart.legend
    if legend is not None:
        legend.txPr = _chart_text_properties(font_name)

    data_labels = chart.dataLabels
    if data_labels is not None:
        data_labels.txPr = _chart_text_properties(font_name)

    x_axis = getattr(chart, "x_axis", None)
    if x_axis is not None:
        x_axis.txPr = _chart_text_properties(font_name)

    y_axis = getattr(chart, "y_axis", None)
    if y_axis is not None:
        y_axis.txPr = _chart_text_properties(font_name)


def _sanitize_sheet_title(value: str, *, fallback: str = "Chart Export") -> str:
    def sanitize(raw_value: str) -> str:
        sanitized_value = "".join(
            character
            for character in raw_value.strip()
            if character not in INVALID_SHEET_TITLE_CHARS
        )
        return sanitized_value[:31].strip()

    sanitized = sanitize(value)
    if sanitized:
        return sanitized

    fallback_title = sanitize(fallback)
    return fallback_title or "Sheet"


def _unique_sheet_title(value: str, *, fallback: str, used_titles: set[str]) -> str:
    base_title = _sanitize_sheet_title(value, fallback=fallback)
    candidate = base_title
    suffix_index = 1

    while candidate in used_titles:
        suffix = f" {suffix_index}"
        trimmed_base_title = base_title[: 31 - len(suffix)].rstrip()
        candidate = f"{trimmed_base_title or fallback[: 31 - len(suffix)]}{suffix}"
        suffix_index += 1

    used_titles.add(candidate)
    return candidate


def _chart_and_data_sheet_titles(title: str) -> tuple[str, str]:
    used_titles: set[str] = set()
    chart_sheet_title = _unique_sheet_title(
        title,
        fallback="Chart Export",
        used_titles=used_titles,
    )
    data_sheet_title = _unique_sheet_title(
        f"{title} Data",
        fallback="Data",
        used_titles=used_titles,
    )
    return chart_sheet_title, data_sheet_title


def _sanitize_table_name(value: str) -> str:
    sanitized = "".join(character for character in value if character.isalnum())
    if not sanitized:
        return "ChartExportData"
    if sanitized[0].isdigit():
        return f"T{sanitized}"
    return sanitized


def _graphical_properties(color: str) -> GraphicalProperties:
    normalized = _normalize_hex_color(color)
    properties = GraphicalProperties(solidFill=normalized)
    properties.line.solidFill = normalized
    return properties


def _chart_color_choice(color: str) -> ColorChoice:
    return ColorChoice(srgbClr=cast(Any, _normalize_hex_color(color)))


def _chart_graphical_properties(
    *,
    fill_color: str | None = None,
    line_color: str | None = None,
    line_width: int | None = None,
    no_fill: bool = False,
    no_line: bool = False,
    round_line: bool = False,
) -> GraphicalProperties:
    properties = GraphicalProperties()

    if fill_color is not None:
        properties.solidFill = _chart_color_choice(fill_color)
    if no_fill:
        properties.noFill = True
    if line_width is not None:
        properties.line.width = line_width
    if line_color is not None:
        properties.line.solidFill = _chart_color_choice(line_color)
    if no_line:
        properties.line.noFill = True
    if round_line:
        properties.line.round = True

    return properties


def _apply_chart_styling(chart: BarChart | PieChart) -> None:
    chart.graphical_properties = _chart_graphical_properties(
        fill_color=CHART_BACKGROUND_HEX,
        line_width=CHART_BORDER_WIDTH,
        no_line=True,
    )
    chart.plot_area.spPr = _chart_graphical_properties(
        fill_color=CHART_BACKGROUND_HEX,
        line_color=CHART_LINE_HEX,
        line_width=AXIS_BORDER_WIDTH,
        round_line=True,
    )

    if not isinstance(chart, BarChart):
        return

    chart.x_axis.spPr = _chart_graphical_properties(
        line_color=CHART_LINE_HEX,
        line_width=AXIS_BORDER_WIDTH,
        round_line=True,
    )
    chart.y_axis.spPr = _chart_graphical_properties(
        line_color=CHART_LINE_HEX,
        line_width=AXIS_BORDER_WIDTH,
        round_line=True,
    )
    chart.y_axis.majorGridlines = ChartLines(
        spPr=_chart_graphical_properties(
            line_color=CHART_LINE_HEX,
            line_width=AXIS_BORDER_WIDTH,
            round_line=True,
        )
    )


def _new_data_point(index: int, color: str) -> Any:
    marker_module = import_module("openpyxl.chart.marker")
    data_point = cast(Any, marker_module.DataPoint)(idx=index)
    data_point.graphicalProperties = _graphical_properties(color)
    return data_point


def _series_reference(
    worksheet: Worksheet,
    min_col: int,
    max_col: int,
    row_count: int,
) -> Reference:
    return Reference(
        worksheet,
        min_col=min_col,
        min_row=HEADER_ROW_INDEX,
        max_col=max_col,
        max_row=HEADER_ROW_INDEX + row_count,
    )


def _category_reference(worksheet: Worksheet, row_count: int) -> Reference:
    return Reference(
        worksheet,
        min_col=1,
        min_row=DATA_ROW_INDEX,
        max_row=DATA_ROW_INDEX + row_count - 1,
    )


def _set_string_categories(
    chart: BarChart | PieChart, worksheet: Worksheet, row_count: int
) -> None:
    category_reference = str(_category_reference(worksheet, row_count))
    for series in chart.ser:
        series.cat = AxDataSource(strRef=StrRef(f=category_reference))


def _configure_data_labels(
    chart: BarChart | PieChart,
    *,
    position: str,
    number_format: str,
    show_percent: bool = False,
) -> None:
    labels = DataLabelList()
    labels.position = position
    labels.numFmt = number_format
    labels.showLegendKey = False
    labels.showCatName = False
    labels.showSerName = False
    if show_percent:
        labels.showVal = False
        labels.showPercent = True
    else:
        labels.showVal = True
    chart.dataLabels = labels


def _set_value_axis_range(
    chart: BarChart,
    minimum: float,
    maximum: float,
    number_format: str,
) -> None:
    chart.y_axis.scaling.min = minimum
    chart.y_axis.scaling.max = maximum
    chart.y_axis.numFmt = number_format


def _reverse_horizontal_category_axis(chart: BarChart) -> None:
    chart.x_axis.scaling.orientation = "maxMin"
    chart.y_axis.crosses = "max"


def _apply_point_fills(series: Any, colors: list[str]) -> None:
    if not colors:
        return

    series.graphicalProperties = _graphical_properties(colors[0])
    series.dPt = [_new_data_point(index, color) for index, color in enumerate(colors)]


def _apply_series_fills(series_list: list[Any], colors: list[str]) -> None:
    for series, color in zip(series_list, colors, strict=True):
        series.graphicalProperties = _graphical_properties(color)


def _style_title_block(
    worksheet: Worksheet,
    title: str,
    meta_line: str,
    column_count: int,
) -> None:
    last_column = max(column_count, 1)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )

    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = _font_with_name(
        CellFont(size=16, bold=True),
        font_name=EXPORT_FONT_NAME,
    )
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    meta_cell = worksheet.cell(row=2, column=1, value=meta_line)
    meta_cell.font = _font_with_name(
        CellFont(size=10, color="FF6B7280"),
        font_name=EXPORT_FONT_NAME,
    )
    meta_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 18


def _style_header_row(
    worksheet: Worksheet,
    spec: WorksheetTableSpec,
) -> None:
    for column_index, header in enumerate(spec.headers, start=1):
        cell = worksheet.cell(row=HEADER_ROW_INDEX, column=column_index, value=header)
        cell.font = _font_with_name(
            CellFont(bold=True),
            font_name=EXPORT_FONT_NAME,
        )
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _set_cell_font_name(cell: Any, *, font_name: str) -> None:
    cast(Any, cell).font = _font_with_name(cast(Any, cell).font, font_name=font_name)


def _apply_table_font(
    worksheet: Worksheet,
    *,
    row_count: int,
    column_count: int,
    font_name: str,
) -> None:
    last_row = HEADER_ROW_INDEX + row_count
    for row in worksheet.iter_rows(
        min_row=HEADER_ROW_INDEX,
        max_row=last_row,
        min_col=1,
        max_col=column_count,
    ):
        for cell in row:
            _set_cell_font_name(cell, font_name=font_name)


def _write_table_rows(worksheet: Worksheet, spec: WorksheetTableSpec) -> None:
    for row_offset, row in enumerate(spec.rows):
        worksheet_row = DATA_ROW_INDEX + row_offset
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=worksheet_row, column=column_index, value=value)
            cell.alignment = Alignment(
                horizontal="right" if isinstance(value, (int, float)) else "left",
                vertical="center",
            )

            fill_color = spec.cell_fill_colors.get((row_offset, column_index - 1))
            if fill_color:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=_argb_hex_color(fill_color),
                )
                cell.font = _font_with_name(
                    CellFont(color=_best_text_color(fill_color)),
                    font_name=EXPORT_FONT_NAME,
                )

            number_format = spec.number_formats.get(column_index - 1)
            if number_format and isinstance(value, (int, float)):
                cell.number_format = number_format


def _set_column_widths(
    worksheet: Worksheet, total_rows: int, column_count: int
) -> None:
    for column_index in range(1, column_count + 1):
        max_length = 0
        for row_index in range(1, total_rows + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            40,
        )


def _add_table(worksheet: Worksheet, spec: WorksheetTableSpec) -> None:
    last_column_letter = get_column_letter(len(spec.headers))
    last_row = HEADER_ROW_INDEX + len(spec.rows)
    table = Table(
        displayName=_sanitize_table_name(spec.table_name),
        ref=f"A{HEADER_ROW_INDEX}:{last_column_letter}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _build_distribution_chart(
    worksheet: Worksheet,
    title: str,
    points: list[dict[str, Any]],
    *,
    horizontal: bool,
) -> BarChart:
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.style = 10
    chart.title = title
    chart.legend = None
    chart.width = 10
    chart.height = max(min(len(points) // 2 + 4, 10), 6)

    chart.add_data(
        _series_reference(worksheet, 2, 2, len(points)), titles_from_data=True
    )
    _set_string_categories(chart, worksheet, len(points))
    _set_value_axis_range(chart, 0.0, 100.0, PERCENT_NUMBER_FORMAT)
    _configure_data_labels(
        chart,
        position="outEnd",
        number_format=PERCENT_NUMBER_FORMAT,
    )
    if horizontal:
        _reverse_horizontal_category_axis(chart)
    _apply_point_fills(chart.ser[0], [point["color"] for point in points])

    return chart


def _build_stacked_bar_chart(
    worksheet: Worksheet,
    title: str,
    rows: list[dict[str, Any]],
) -> BarChart:
    if not rows:
        raise ValueError("Stacked chart export requires at least one row")

    chart = BarChart()
    chart.type = "bar"
    chart.style = 13
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    legend = chart.legend
    if legend is not None:
        legend.position = "b"

    chart.width = 10
    chart.height = max(min(len(rows) // 2 + 4, 10), 6)

    chart.add_data(
        _series_reference(worksheet, 2, len(rows[0]["segments"]) + 1, len(rows)),
        titles_from_data=True,
    )
    _set_string_categories(chart, worksheet, len(rows))
    _set_value_axis_range(chart, 0.0, 100.0, STACKED_CHART_NUMBER_FORMAT)
    _configure_data_labels(
        chart,
        position="ctr",
        number_format=STACKED_CHART_NUMBER_FORMAT,
    )
    _reverse_horizontal_category_axis(chart)
    _apply_series_fills(
        list(chart.ser),
        [segment["color"] for segment in rows[0]["segments"]],
    )

    return chart


def _build_pie_chart(
    worksheet: Worksheet,
    title: str,
    points: list[dict[str, Any]],
) -> PieChart:
    chart = PieChart()
    chart.title = title
    legend = chart.legend
    if legend is not None:
        legend.position = "b"

    chart.width = 10
    chart.height = 7

    chart.add_data(
        _series_reference(worksheet, 2, 2, len(points)), titles_from_data=True
    )
    _set_string_categories(chart, worksheet, len(points))
    _configure_data_labels(
        chart,
        position="outEnd",
        number_format=PIE_PERCENT_NUMBER_FORMAT,
        show_percent=True,
    )
    chart.series[0].data_points = [
        _new_data_point(index, point["color"]) for index, point in enumerate(points)
    ]

    return chart


def _build_mean_bar_chart(
    worksheet: Worksheet,
    title: str,
    chart_payload: dict[str, Any],
) -> BarChart:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = title
    chart.legend = None
    chart.width = 10
    chart.height = 6

    chart.add_data(
        _series_reference(worksheet, 2, 2, len(chart_payload["points"])),
        titles_from_data=True,
    )
    _set_string_categories(chart, worksheet, len(chart_payload["points"]))
    _set_value_axis_range(
        chart,
        float(chart_payload["min_value"]),
        float(chart_payload["max_value"]),
        MEAN_NUMBER_FORMAT,
    )
    _configure_data_labels(
        chart,
        position="outEnd",
        number_format=MEAN_NUMBER_FORMAT,
    )
    _reverse_horizontal_category_axis(chart)
    _apply_series_fills(list(chart.ser), [chart_payload["color"]])

    return chart


def _build_distribution_table_spec(
    chart: dict[str, Any], labels: dict[str, str], table_name: str
) -> WorksheetTableSpec:
    rows = [[point["label"], point["value"]] for point in chart["points"]]

    return WorksheetTableSpec(
        headers=[labels["label"], labels["value_percent"]],
        rows=rows,
        number_formats={1: PERCENT_NUMBER_FORMAT},
        table_name=table_name,
    )


def _build_stacked_table_spec(
    chart: dict[str, Any], labels: dict[str, str]
) -> WorksheetTableSpec:
    first_row = chart["rows"][0]
    headers = [
        labels["label"],
        *[segment["label"] for segment in first_row["segments"]],
    ]
    rows = [
        [row["label"], *[segment["value"] for segment in row["segments"]]]
        for row in chart["rows"]
    ]
    header_fill_colors = {
        index + 1: segment["color"]
        for index, segment in enumerate(first_row["segments"])
    }
    number_formats = cast(
        dict[int, str],
        dict.fromkeys(range(1, len(headers)), STACKED_CHART_NUMBER_FORMAT),
    )

    return WorksheetTableSpec(
        headers=headers,
        rows=rows,
        header_fill_colors=header_fill_colors,
        number_formats=number_formats,
        table_name="StackedChartData",
    )


def _build_mean_bar_table_spec(
    chart: dict[str, Any], labels: dict[str, str]
) -> WorksheetTableSpec:
    return WorksheetTableSpec(
        headers=[labels["metric"], labels["value"]],
        rows=[[point["label"], point["value"]] for point in chart["points"]],
        header_fill_colors={1: chart["color"]},
        number_formats={1: MEAN_NUMBER_FORMAT},
        table_name="MeanBarData",
    )


def _build_metrics_table_spec(
    chart: dict[str, Any], labels: dict[str, str]
) -> WorksheetTableSpec:
    return WorksheetTableSpec(
        headers=[labels["metric"], labels["value"]],
        rows=[[metric["label"], metric["value"]] for metric in chart["metrics"]],
        table_name="MetricsData",
    )


def _build_table_spec(
    chart: dict[str, Any], labels: dict[str, str]
) -> WorksheetTableSpec:
    kind = chart["kind"]

    if kind in {"bar", "horizontalBar", "multiResponse"}:
        return _build_distribution_table_spec(chart, labels, "DistributionChartData")
    if kind == "horizontalStackedBar":
        return _build_stacked_table_spec(chart, labels)
    if kind == "pie":
        return _build_distribution_table_spec(chart, labels, "PieChartData")
    if kind == "meanBar":
        return _build_mean_bar_table_spec(chart, labels)
    if kind == "metrics":
        return _build_metrics_table_spec(chart, labels)

    raise ValueError(f"Unsupported chart kind '{kind}'")


def _build_chart_for_payload(
    worksheet: Worksheet,
    title: str,
    chart_payload: dict[str, Any],
) -> BarChart | PieChart | None:
    kind = chart_payload["kind"]

    if kind == "bar":
        return _build_distribution_chart(
            worksheet,
            title,
            chart_payload["points"],
            horizontal=False,
        )
    if kind in {"horizontalBar", "multiResponse"}:
        return _build_distribution_chart(
            worksheet,
            title,
            chart_payload["points"],
            horizontal=True,
        )
    if kind == "horizontalStackedBar":
        return _build_stacked_bar_chart(
            worksheet,
            title,
            chart_payload["rows"],
        )
    if kind == "pie":
        return _build_pie_chart(
            worksheet,
            title,
            chart_payload["points"],
        )
    if kind == "meanBar":
        return _build_mean_bar_chart(
            worksheet,
            title,
            chart_payload,
        )
    if kind == "metrics":
        return None

    raise ValueError(f"Unsupported chart kind '{kind}'")


def _add_chart_sheet(
    workbook: Workbook,
    *,
    chart: BarChart | PieChart,
    title: str,
) -> None:
    chartsheet = workbook.create_chartsheet(title=title, index=0)
    chartsheet.add_chart(chart)
    cast(Any, chart).anchor = AbsoluteAnchor(
        pos=XDRPoint2D(0, 0),
        ext=XDRPositiveSize2D(
            cx=cm_to_EMU(chart.width * CHARTSHEET_SIZE_MULTIPLIER),
            cy=cm_to_EMU(chart.height * CHARTSHEET_SIZE_MULTIPLIER),
        ),
    )
    chartsheet.sheetViews.sheetView[0].zoomToFit = True
    workbook.active = chartsheet


def build_workbook(payload: dict[str, Any]) -> bytes:
    """Build an Excel workbook from the normalized export payload."""
    workbook = Workbook()
    _set_workbook_default_font(workbook, font_name=EXPORT_FONT_NAME)
    worksheet = cast(Worksheet, workbook.active)
    chart_sheet_title: str | None = None
    if payload["chart"]["kind"] == "metrics":
        worksheet.title = _sanitize_sheet_title(payload["title"])
    else:
        chart_sheet_title, data_sheet_title = _chart_and_data_sheet_titles(
            payload["title"]
        )
        worksheet.title = data_sheet_title

    worksheet.freeze_panes = f"A{DATA_ROW_INDEX}"
    worksheet.sheet_view.showGridLines = True

    spec = _build_table_spec(payload["chart"], payload["labels"])

    _style_title_block(
        worksheet, payload["title"], payload["meta_line"], len(spec.headers)
    )
    _style_header_row(worksheet, spec)
    _write_table_rows(worksheet, spec)
    _add_table(worksheet, spec)
    _apply_table_font(
        worksheet,
        row_count=len(spec.rows),
        column_count=len(spec.headers),
        font_name=EXPORT_FONT_NAME,
    )
    _set_column_widths(
        worksheet,
        total_rows=HEADER_ROW_INDEX + len(spec.rows),
        column_count=len(spec.headers),
    )
    chart = _build_chart_for_payload(
        worksheet,
        payload["title"],
        payload["chart"],
    )
    if chart is not None:
        _apply_chart_styling(chart)
        _apply_chart_font(chart, font_name=EXPORT_FONT_NAME)
    if chart is not None and chart_sheet_title is not None:
        _add_chart_sheet(workbook, chart=chart, title=chart_sheet_title)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_content_disposition(file_name: str) -> str:
    """Create a safe attachment header value for the generated workbook."""
    safe_name = (
        Path(file_name).name.replace("\r", "").replace("\n", "").replace('"', "")
    )
    if not safe_name:
        safe_name = "chart-export.xlsx"
    return f'attachment; filename="{safe_name}"'
