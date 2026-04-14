import re
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from openpyxl.chartsheet.chartsheet import Chartsheet
from pydantic import ValidationError

from analysis.services.excel_export import EXPORT_FONT_NAME, build_workbook
from analysis.web.api.schemas.datasets import ExcelExportRequest


def _chart_xml(workbook_bytes: bytes) -> str:
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        return archive.read("xl/charts/chart1.xml").decode("utf-8")


def test_build_workbook_for_mean_bar_uses_numeric_axis_format() -> None:
    """Mean bar exports should keep a numeric axis instead of percentages."""
    payload = {
        "file_name": "mean-bar-2026-03-17.xlsx",
        "title": "Satisfaction",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#3b82f6"],
        "chart": {
            "kind": "meanBar",
            "points": [
                {"label": "Mean", "value": 4.2},
                {"label": "Median", "value": 4.0},
            ],
            "min_value": 1,
            "max_value": 5,
            "color": "#3b82f6",
        },
    }

    workbook_bytes = build_workbook(payload)

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert 'formatCode="0.0"' in chart_xml
    assert '<max val="5"/>' in chart_xml
    assert '<min val="1"/>' in chart_xml


def test_build_workbook_for_distribution_preserves_point_colors() -> None:
    """Distribution exports should keep each point color in the chart."""
    payload = {
        "file_name": "bar-export-2026-03-17.xlsx",
        "title": "Age group",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    workbook_bytes = build_workbook(payload)

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert '<a:srgbClr val="FF0000"/>' in chart_xml
    assert '<a:srgbClr val="00FF00"/>' in chart_xml


def test_build_workbook_for_distribution_uses_whole_percent_axis_and_labels() -> None:
    """Distribution bar exports should show whole-percent values in chart labels and axes."""
    payload = {
        "file_name": "bar-export-2026-03-17.xlsx",
        "title": "Age group",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    chart_xml = _chart_xml(build_workbook(payload))

    assert 'formatCode="0&quot;%&quot;" sourceLinked="0"' in chart_xml
    assert 'formatCode="0&quot;%&quot;"' in chart_xml
    assert 'formatCode="0.00&quot;%&quot;"' not in chart_xml


def test_build_workbook_for_bar_chart_uses_light_gray_axis_and_plot_area_lines() -> (
    None
):
    """Bar chart exports should use the sandbox light-gray axis and plot-area lines."""
    payload = {
        "file_name": "bar-export-2026-03-24.xlsx",
        "title": "Age group",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 24, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    chart_xml = _chart_xml(build_workbook(payload))

    assert re.search(
        r"<catAx>.*?<spPr><a:ln[^>]*><a:solidFill><a:srgbClr val=\"999999\"/></a:solidFill>",
        chart_xml,
    )
    assert re.search(
        r"<valAx>.*?<majorGridlines><spPr><a:ln[^>]*><a:solidFill><a:srgbClr val=\"999999\"/></a:solidFill>",
        chart_xml,
    )
    assert re.search(
        r"<valAx>.*?<spPr><a:ln[^>]*><a:solidFill><a:srgbClr val=\"999999\"/></a:solidFill>",
        chart_xml,
    )
    assert (
        "<plotArea>" in chart_xml
        and '<a:srgbClr val="FFFFFF"/>' in chart_xml
        and '<a:srgbClr val="999999"/>' in chart_xml
    )


def test_build_workbook_for_pie_chart_uses_light_gray_plot_area_line() -> None:
    """Pie chart exports should use the same light-gray plot-area border."""
    payload = {
        "file_name": "pie-export-2026-03-24.xlsx",
        "title": "Preferred contact",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 24, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "pie",
            "points": [
                {"label": "Email", "value": 55, "color": "#ff0000"},
                {"label": "Phone", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    chart_xml = _chart_xml(build_workbook(payload))

    assert re.search(
        r"<plotArea>.*?<spPr><a:solidFill[^>]*><a:srgbClr val=\"FFFFFF\"/></a:solidFill><a:ln[^>]*><a:solidFill><a:srgbClr val=\"999999\"/></a:solidFill>",
        chart_xml,
    )


@pytest.mark.parametrize(
    "chart",
    [
        {
            "kind": "horizontalBar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
        {
            "kind": "multiResponse",
            "points": [
                {"label": "Email", "value": 55, "color": "#ff0000"},
                {"label": "Phone", "value": 45, "color": "#00ff00"},
            ],
        },
        {
            "kind": "horizontalStackedBar",
            "rows": [
                {
                    "label": "Group A",
                    "segments": [
                        {"label": "Yes", "value": 60, "color": "#1f2937"},
                        {"label": "No", "value": 40, "color": "#e5e7eb"},
                    ],
                }
            ],
        },
        {
            "kind": "meanBar",
            "points": [
                {"label": "Mean", "value": 4.2},
                {"label": "Median", "value": 4.0},
            ],
            "min_value": 1,
            "max_value": 5,
            "color": "#3b82f6",
        },
    ],
    ids=["horizontal-bar", "multi-response", "stacked", "mean-bar"],
)
def test_build_workbook_for_horizontal_charts_reverses_category_axis_and_axis_crossing(
    chart: dict[str, object],
) -> None:
    """Horizontal charts should match worksheet row order and keep the axis line at the end."""
    payload = {
        "file_name": f"{chart['kind']}-export-2026-03-24.xlsx",
        "title": "Chart export",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 24, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#3b82f6", "#ef4444"],
        "chart": chart,
    }

    workbook_bytes = build_workbook(payload)

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert re.search(
        r"<catAx>.*?<orientation val=\"maxMin\"/>.*?</catAx>",
        chart_xml,
    )
    assert re.search(r"<valAx>.*?<crosses val=\"max\"/>.*?</valAx>", chart_xml)
    assert not re.search(r"<catAx>.*?<crosses val=\"max\"/>.*?</catAx>", chart_xml)


@pytest.mark.parametrize(
    ("chart", "expected_chart_size"),
    [
        (
            {
                "kind": "bar",
                "points": [
                    {"label": "18-29", "value": 55, "color": "#ff0000"},
                    {"label": "30-44", "value": 45, "color": "#00ff00"},
                ],
            },
            (7200000, 4320000),
        ),
        (
            {
                "kind": "horizontalBar",
                "points": [
                    {"label": "18-29", "value": 55, "color": "#ff0000"},
                    {"label": "30-44", "value": 45, "color": "#00ff00"},
                ],
            },
            (7200000, 4320000),
        ),
        (
            {
                "kind": "multiResponse",
                "points": [
                    {"label": "Email", "value": 55, "color": "#ff0000"},
                    {"label": "Phone", "value": 45, "color": "#00ff00"},
                ],
            },
            (7200000, 4320000),
        ),
        (
            {
                "kind": "horizontalStackedBar",
                "rows": [
                    {
                        "label": "Group A",
                        "segments": [
                            {"label": "Yes", "value": 60, "color": "#1f2937"},
                            {"label": "No", "value": 40, "color": "#e5e7eb"},
                        ],
                    }
                ],
            },
            (7200000, 4320000),
        ),
        (
            {
                "kind": "pie",
                "points": [
                    {"label": "Yes", "value": 55, "color": "#ff0000"},
                    {"label": "No", "value": 45, "color": "#00ff00"},
                ],
            },
            (7200000, 5040000),
        ),
        (
            {
                "kind": "meanBar",
                "points": [
                    {"label": "Mean", "value": 4.2},
                    {"label": "Median", "value": 4.0},
                ],
                "min_value": 1,
                "max_value": 5,
                "color": "#3b82f6",
            },
            (7200000, 4320000),
        ),
    ],
    ids=["bar", "horizontal-bar", "multi-response", "stacked", "pie", "mean-bar"],
)
def test_build_workbook_for_chart_exports_opens_on_chartsheet(
    chart: dict[str, object],
    expected_chart_size: tuple[int, int],
) -> None:
    """Chart exports should open on a chartsheet while keeping data on a worksheet."""
    payload = {
        "file_name": f"{chart['kind']}-export-2026-03-17.xlsx",
        "title": "Chart export",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#3b82f6", "#ef4444"],
        "chart": chart,
    }

    workbook_bytes = build_workbook(payload)
    workbook = load_workbook(BytesIO(workbook_bytes))

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chartsheet_entries = [
            name
            for name in archive.namelist()
            if name.startswith("xl/chartsheets/sheet") and name.endswith(".xml")
        ]
        drawing_xml = archive.read("xl/drawings/drawing1.xml").decode("utf-8")

    assert isinstance(workbook.active, Chartsheet)
    assert chartsheet_entries == ["xl/chartsheets/sheet1.xml"]
    assert 'cx="0"' not in drawing_xml
    assert 'cy="0"' not in drawing_xml
    assert f'cx="{expected_chart_size[0]}"' in drawing_xml
    assert f'cy="{expected_chart_size[1]}"' in drawing_xml
    assert len(workbook.chartsheets) == 1
    assert len(workbook.worksheets) == 1
    assert workbook.sheetnames[0] == "Chart export"
    assert workbook.sheetnames[1] == "Chart export Data"
    assert workbook.active.title == "Chart export"
    assert workbook.worksheets[0].title == "Chart export Data"


@pytest.mark.parametrize(
    ("title", "expected_sheetnames", "expected_active_title"),
    [
        (
            "1234567890123456789012345678901",
            ["1234567890123456789012345678901", "12345678901234567890123456789 1"],
            "1234567890123456789012345678901",
        ),
        (
            "[]:*?/\\",
            ["Chart Export", "Data"],
            "Chart Export",
        ),
    ],
    ids=["truncated-title-collision", "invalid-title-fallback"],
)
def test_build_workbook_generates_unique_valid_sheet_titles(
    title: str,
    expected_sheetnames: list[str],
    expected_active_title: str,
) -> None:
    """Chart exports should create distinct valid worksheet and chartsheet titles."""
    payload = {
        "file_name": "chart-export-2026-03-17.xlsx",
        "title": title,
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    workbook = load_workbook(BytesIO(build_workbook(payload)))

    assert workbook.sheetnames == expected_sheetnames
    assert isinstance(workbook.active, Chartsheet)
    assert workbook.active.title == expected_active_title
    assert all(len(sheet_name) <= 31 for sheet_name in workbook.sheetnames)
    assert len(set(workbook.sheetnames)) == len(workbook.sheetnames)


@pytest.mark.parametrize(
    ("chart", "table_name"),
    [
        (
            {
                "kind": "bar",
                "points": [
                    {"label": "18-29", "value": 55, "color": "#ff0000"},
                    {"label": "30-44", "value": 45, "color": "#00ff00"},
                ],
            },
            "DistributionChartData",
        ),
        (
            {
                "kind": "horizontalBar",
                "points": [
                    {"label": "18-29", "value": 55, "color": "#ff0000"},
                    {"label": "30-44", "value": 45, "color": "#00ff00"},
                ],
            },
            "DistributionChartData",
        ),
        (
            {
                "kind": "multiResponse",
                "points": [
                    {"label": "Email", "value": 55, "color": "#ff0000"},
                    {"label": "Phone", "value": 45, "color": "#00ff00"},
                ],
            },
            "DistributionChartData",
        ),
        (
            {
                "kind": "pie",
                "points": [
                    {"label": "Yes", "value": 55, "color": "#ff0000"},
                    {"label": "No", "value": 45, "color": "#00ff00"},
                ],
            },
            "PieChartData",
        ),
    ],
    ids=["bar", "horizontal-bar", "multi-response", "pie"],
)
def test_build_workbook_omits_color_column_for_point_based_charts(
    chart: dict[str, object], table_name: str
) -> None:
    """Point-based chart exports should not include a worksheet color column."""
    payload = {
        "file_name": f"{chart['kind']}-export-2026-03-17.xlsx",
        "title": "Chart export",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": chart,
    }

    workbook = load_workbook(BytesIO(build_workbook(payload)))
    worksheet = workbook.worksheets[0]

    assert isinstance(workbook.active, Chartsheet)
    assert worksheet.tables[table_name].ref == "A4:B6"
    assert worksheet["A4"].value == "Label"
    assert worksheet["B4"].value == "Value (%)"
    assert worksheet["C4"].value is None
    assert worksheet["C5"].value is None
    assert worksheet["C6"].value is None


def test_build_workbook_for_stacked_chart_uses_percent_format_and_legend() -> None:
    """Stacked chart exports should use whole-percent labels and a bottom legend."""
    payload = {
        "file_name": "stacked-export-2026-03-17.xlsx",
        "title": "Awareness by segment",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#1f2937", "#e5e7eb"],
        "chart": {
            "kind": "horizontalStackedBar",
            "rows": [
                {
                    "label": "Group A",
                    "segments": [
                        {"label": "Yes", "value": 60, "color": "#1f2937"},
                        {"label": "No", "value": 40, "color": "#e5e7eb"},
                    ],
                }
            ],
        },
    }

    workbook_bytes = build_workbook(payload)

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert 'grouping val="stacked"' in chart_xml
    assert 'legendPos val="b"' in chart_xml
    assert 'formatCode="0&quot;%&quot;" sourceLinked="0"' in chart_xml
    assert 'formatCode="0&quot;%&quot;"' in chart_xml
    assert '<max val="100"/>' in chart_xml
    assert '<min val="0"/>' in chart_xml
    assert '<showVal val="1"/>' in chart_xml
    assert '<showCatName val="0"/>' in chart_xml
    assert '<showSerName val="0"/>' in chart_xml
    assert '<showPercent val="1"/>' not in chart_xml
    assert "<cat><strRef>" in chart_xml
    assert "<cat><numRef>" not in chart_xml


def test_build_workbook_for_stacked_chart_formats_source_cells_as_whole_percent() -> (
    None
):
    """Stacked chart worksheet cells should use whole-percent formatting."""
    payload = {
        "file_name": "stacked-export-2026-03-17.xlsx",
        "title": "Awareness by segment",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#1f2937", "#e5e7eb"],
        "chart": {
            "kind": "horizontalStackedBar",
            "rows": [
                {
                    "label": "Group A",
                    "segments": [
                        {"label": "White", "value": 79.14, "color": "#1f2937"},
                        {"label": "Black", "value": 14.12, "color": "#e5e7eb"},
                        {"label": "Other", "value": 6.74, "color": "#94a3b8"},
                    ],
                }
            ],
        },
    }

    workbook = load_workbook(BytesIO(build_workbook(payload)))
    worksheet = workbook.worksheets[0]

    assert isinstance(workbook.active, Chartsheet)
    assert worksheet["B5"].number_format == '0"%"'
    assert worksheet["C5"].number_format == '0"%"'
    assert worksheet["D5"].number_format == '0"%"'
    assert worksheet["B5"].value == 79.14
    assert worksheet["C5"].value == 14.12
    assert worksheet["D5"].value == 6.74


def test_build_workbook_for_distribution_formats_source_cells_as_whole_percent() -> (
    None
):
    """Distribution chart worksheet cells should use whole-percent formatting."""
    payload = {
        "file_name": "bar-export-2026-03-17.xlsx",
        "title": "Age group",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    workbook = load_workbook(BytesIO(build_workbook(payload)))
    worksheet = workbook.worksheets[0]

    assert isinstance(workbook.active, Chartsheet)
    assert worksheet["B5"].number_format == '0"%"'
    assert worksheet["B6"].number_format == '0"%"'
    assert worksheet["B5"].value == 55
    assert worksheet["B6"].value == 45


def test_build_workbook_for_metrics_creates_styled_table_without_chart() -> None:
    """Metrics exports should generate a workbook table with no chart sheet parts."""
    payload = {
        "file_name": "metrics-2026-03-17.xlsx",
        "title": "Metrics",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#3b82f6", "#ef4444"],
        "chart": {
            "kind": "metrics",
            "metrics": [
                {"label": "Count", "value": "100"},
                {"label": "Mean", "value": "1.5"},
            ],
        },
    }

    workbook_bytes = build_workbook(payload)
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook.active

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        chartsheet_entries = [
            name for name in archive.namelist() if name.startswith("xl/chartsheets/")
        ]

    assert worksheet is not None
    assert not chartsheet_entries
    assert worksheet["A5"].value == "Count"
    assert worksheet["B6"].value == "1.5"
    assert worksheet.tables["MetricsData"].ref == "A4:B6"


@pytest.mark.parametrize(
    ("payload", "table_cells"),
    [
        (
            {
                "file_name": "bar-export-2026-03-17.xlsx",
                "title": "Age group",
                "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
                "labels": {
                    "label": "Label",
                    "value": "Value",
                    "value_percent": "Value (%)",
                    "color": "Color",
                    "metric": "Metric",
                },
                "palette": ["#ff0000", "#00ff00"],
                "chart": {
                    "kind": "bar",
                    "points": [
                        {"label": "18-29", "value": 55, "color": "#ff0000"},
                        {"label": "30-44", "value": 45, "color": "#00ff00"},
                    ],
                },
            },
            ["A4", "B4", "A5", "B6"],
        ),
        (
            {
                "file_name": "metrics-2026-03-17.xlsx",
                "title": "Metrics",
                "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
                "labels": {
                    "label": "Label",
                    "value": "Value",
                    "value_percent": "Value (%)",
                    "color": "Color",
                    "metric": "Metric",
                },
                "palette": ["#3b82f6", "#ef4444"],
                "chart": {
                    "kind": "metrics",
                    "metrics": [
                        {"label": "Count", "value": "100"},
                        {"label": "Mean", "value": "1.5"},
                    ],
                },
            },
            ["A4", "B4", "A5", "B6"],
        ),
    ],
    ids=["chart-table", "metrics-table"],
)
def test_build_workbook_uses_sans_serif_font_for_all_table_cells(
    payload: dict[str, object],
    table_cells: list[str],
) -> None:
    """Excel table headers and data cells should use a consistent sans-serif font."""
    workbook = load_workbook(BytesIO(build_workbook(payload)))
    worksheet = workbook.worksheets[0]

    assert worksheet["A4"].font.bold
    for cell_reference in table_cells:
        assert worksheet[cell_reference].font.name == EXPORT_FONT_NAME


def test_build_workbook_uses_same_font_for_title_meta_and_chart_text() -> None:
    """Worksheet and chart text should use the same sans-serif font family."""
    payload = {
        "file_name": "bar-export-2026-03-17.xlsx",
        "title": "Age group",
        "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
        "labels": {
            "label": "Label",
            "value": "Value",
            "value_percent": "Value (%)",
            "color": "Color",
            "metric": "Metric",
        },
        "palette": ["#ff0000", "#00ff00"],
        "chart": {
            "kind": "bar",
            "points": [
                {"label": "18-29", "value": 55, "color": "#ff0000"},
                {"label": "30-44", "value": 45, "color": "#00ff00"},
            ],
        },
    }

    workbook_bytes = build_workbook(payload)
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook.worksheets[0]

    with ZipFile(BytesIO(workbook_bytes)) as archive:
        styles_xml = archive.read("xl/styles.xml").decode("utf-8")
        chart_xml = archive.read("xl/charts/chart1.xml").decode("utf-8")

    assert worksheet["A1"].font.name == EXPORT_FONT_NAME
    assert worksheet["A2"].font.name == EXPORT_FONT_NAME
    assert 'name val="Calibri"' not in styles_xml
    assert f'typeface="{EXPORT_FONT_NAME}"' in chart_xml


def test_excel_request_model_accepts_bar_export() -> None:
    """Excel export request validation should allow valid chart payloads."""
    request = ExcelExportRequest.model_validate(
        {
            "file_name": "chart-export-2026-03-17.xlsx",
            "title": "Age group",
            "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
            "labels": {
                "label": "Label",
                "value": "Value",
                "value_percent": "Value (%)",
                "color": "Color",
                "metric": "Metric",
            },
            "palette": ["#3b82f6", "#ef4444"],
            "chart": {
                "kind": "bar",
                "points": [
                    {"label": "18-29", "value": 55.0, "color": "#3b82f6"},
                    {"label": "30-44", "value": 45.0, "color": "#3b82f6"},
                ],
            },
        }
    )

    assert request.chart.kind == "bar"
    assert request.file_name.endswith(".xlsx")


def test_excel_request_rejects_too_many_metric_cards() -> None:
    """Metrics export is limited to six rows to match the existing export shape."""
    with pytest.raises(ValidationError):
        ExcelExportRequest.model_validate(
            {
                "file_name": "metrics-2026-03-17.xlsx",
                "title": "Metrics",
                "meta_line": "Dataset: Survey 2026 | Exported: Mar 17, 2026",
                "labels": {
                    "label": "Label",
                    "value": "Value",
                    "value_percent": "Value (%)",
                    "color": "Color",
                    "metric": "Metric",
                },
                "palette": ["#3b82f6"],
                "chart": {
                    "kind": "metrics",
                    "metrics": [
                        {"label": f"Metric {index}", "value": str(index)}
                        for index in range(7)
                    ],
                },
            }
        )


def test_excel_request_rejects_inconsistent_stacked_rows() -> None:
    """Stacked chart rows must share the same segment order."""
    with pytest.raises(ValidationError):
        ExcelExportRequest.model_validate(
            {
                "file_name": "stacked-2026-03-17.xlsx",
                "title": "Age by gender",
                "meta_line": "Dataset: Survey 2026 | Split: Gender | Exported: Mar 17, 2026",
                "labels": {
                    "label": "Label",
                    "value": "Value",
                    "value_percent": "Value (%)",
                    "color": "Color",
                    "metric": "Metric",
                },
                "palette": ["#3b82f6", "#ef4444"],
                "chart": {
                    "kind": "horizontalStackedBar",
                    "rows": [
                        {
                            "label": "Women",
                            "segments": [
                                {"label": "Yes", "value": 60, "color": "#3b82f6"},
                                {"label": "No", "value": 40, "color": "#ef4444"},
                            ],
                        },
                        {
                            "label": "Men",
                            "segments": [
                                {"label": "No", "value": 45, "color": "#ef4444"},
                                {"label": "Yes", "value": 55, "color": "#3b82f6"},
                            ],
                        },
                    ],
                },
            }
        )


def test_excel_request_rejects_inconsistent_stacked_row_colors() -> None:
    """Stacked chart rows must share the same colors for matching segments."""
    with pytest.raises(ValidationError):
        ExcelExportRequest.model_validate(
            {
                "file_name": "stacked-colors-2026-03-17.xlsx",
                "title": "Age by gender",
                "meta_line": "Dataset: Survey 2026 | Split: Gender | Exported: Mar 17, 2026",
                "labels": {
                    "label": "Label",
                    "value": "Value",
                    "value_percent": "Value (%)",
                    "color": "Color",
                    "metric": "Metric",
                },
                "palette": ["#3b82f6", "#ef4444"],
                "chart": {
                    "kind": "horizontalStackedBar",
                    "rows": [
                        {
                            "label": "Women",
                            "segments": [
                                {"label": "Yes", "value": 60, "color": "#3b82f6"},
                                {"label": "No", "value": 40, "color": "#ef4444"},
                            ],
                        },
                        {
                            "label": "Men",
                            "segments": [
                                {"label": "Yes", "value": 55, "color": "#22c55e"},
                                {"label": "No", "value": 45, "color": "#ef4444"},
                            ],
                        },
                    ],
                },
            }
        )


def test_build_workbook_uses_localized_headers() -> None:
    """Excel exports should use localized worksheet headers from the payload."""
    payload = {
        "file_name": "metrics-2026-03-17.xlsx",
        "title": "Kennzahlen",
        "meta_line": "Datensatz: Umfrage 2026 | Exportiert: 17. Mar. 2026",
        "labels": {
            "label": "Bezeichnung",
            "value": "Wert",
            "value_percent": "Wert (%)",
            "color": "Farbe",
            "metric": "Kennzahl",
        },
        "palette": ["#3b82f6"],
        "chart": {
            "kind": "metrics",
            "metrics": [
                {"label": "Anzahl", "value": "100"},
            ],
        },
    }

    workbook = load_workbook(BytesIO(build_workbook(payload)))
    worksheet = workbook.active

    assert worksheet is not None
    assert worksheet["A4"].value == "Kennzahl"
    assert worksheet["B4"].value == "Wert"
